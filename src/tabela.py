"""Tabela de dados leve — o suficiente para substituir o pandas na interface.

O app usava DataFrame só como "lista de linhas com colunas nomeadas": ler uma
consulta SQL, filtrar por texto, ordenar e percorrer. Nada de estatística, de
séries temporais nem de álgebra — mas pandas e numpy custavam ~27 MB no APK
(mais que a soma de todo o resto do Python empacotado), o que pesava na
instalação pelo celular.

Aqui ficam só as operações realmente usadas, com os mesmos nomes de antes
(``empty``, ``columns``, ``itertuples``, ``iloc``, ``to_dict``,
``sort_values``) para os pontos de uso continuarem legíveis para quem conhecia
o código com pandas.

Uma ``Tabela`` é imutável na prática: filtrar e ordenar devolvem uma nova.
"""

from __future__ import annotations

import sqlite3
from typing import Any, Iterable, Iterator, NamedTuple


class Mascara:
    """Vetor de booleanos para filtrar uma Tabela (``tabela[mascara]``)."""

    __slots__ = ("valores",)

    def __init__(self, valores: Iterable[bool]) -> None:
        self.valores = [bool(v) for v in valores]

    def __or__(self, outra: Mascara) -> Mascara:
        return Mascara(a or b for a, b in zip(self.valores, outra.valores))

    def __and__(self, outra: Mascara) -> Mascara:
        return Mascara(a and b for a, b in zip(self.valores, outra.valores))

    def __invert__(self) -> Mascara:
        return Mascara(not v for v in self.valores)

    def __iter__(self) -> Iterator[bool]:
        return iter(self.valores)

    def __len__(self) -> int:
        return len(self.valores)

    def sum(self) -> int:
        return sum(1 for v in self.valores if v)


class Coluna:
    """Uma coluna: itera como lista e compara como vetor (vira ``Mascara``)."""

    __slots__ = ("valores",)

    def __init__(self, valores: Iterable[Any]) -> None:
        self.valores = list(valores)

    def __iter__(self) -> Iterator[Any]:
        return iter(self.valores)

    def __len__(self) -> int:
        return len(self.valores)

    def __getitem__(self, indice: int) -> Any:
        return self.valores[indice]

    @property
    def iloc(self) -> list[Any]:
        """``coluna.iloc[0]`` — mesmo acesso por posição que a Tabela tem."""
        return self.valores

    def __eq__(self, outro: Any) -> Mascara:  # type: ignore[override]
        return Mascara(v == outro for v in self.valores)

    def __ne__(self, outro: Any) -> Mascara:  # type: ignore[override]
        return Mascara(v != outro for v in self.valores)

    def astype(self, tipo: type) -> Coluna:
        """Converte os valores; nulo vira string vazia em vez de "None"."""
        if tipo is str:
            return Coluna("" if v is None else str(v) for v in self.valores)
        return Coluna(tipo(v) for v in self.valores)

    def dropna(self) -> Coluna:
        return Coluna(v for v in self.valores if v is not None)

    def unique(self) -> list[Any]:
        vistos: list[Any] = []
        for valor in self.valores:
            if valor not in vistos:
                vistos.append(valor)
        return vistos

    def nao_comeca_com(self, prefixo: str) -> Mascara:
        """Máscara do que NÃO começa com o prefixo (nulo conta como "não")."""
        return Mascara(
            not ("" if v is None else str(v)).startswith(prefixo) for v in self.valores
        )

    def contem(self, termo: str) -> Mascara:
        """Máscara de "contém este texto", sem diferenciar maiúsculas."""
        alvo = termo.lower()
        return Mascara(
            alvo in ("" if v is None else str(v)).lower() for v in self.valores
        )


class Tabela:
    """Linhas (dicionários) com uma ordem de colunas fixa."""

    __slots__ = ("linhas", "colunas")

    def __init__(self, linhas: list[dict], colunas: list[str] | None = None) -> None:
        self.linhas = linhas
        if colunas is not None:
            self.colunas = list(colunas)
        else:
            # Sem colunas declaradas (ex.: lista de dicionários vinda de fora),
            # a ordem é a da primeira linha.
            self.colunas = list(linhas[0].keys()) if linhas else []

    # -- construção ---------------------------------------------------------

    @classmethod
    def de_consulta(cls, conn: sqlite3.Connection, sql: str, params=None) -> Tabela:
        """Roda uma consulta e devolve o resultado (era ``pd.read_sql_query``)."""
        cursor = conn.execute(sql, params or ())
        colunas = [descricao[0] for descricao in cursor.description]
        return cls([dict(zip(colunas, linha)) for linha in cursor.fetchall()], colunas)

    # -- leitura ------------------------------------------------------------

    @property
    def empty(self) -> bool:
        return not self.linhas

    def __len__(self) -> int:
        return len(self.linhas)

    def __iter__(self) -> Iterator[dict]:
        return iter(self.linhas)

    def __getitem__(self, chave):
        """``tabela["coluna"]`` devolve a coluna; ``tabela[mascara]`` filtra."""
        if isinstance(chave, Mascara):
            return Tabela(
                [linha for linha, manter in zip(self.linhas, chave) if manter],
                self.colunas,
            )
        return Coluna(linha.get(chave) for linha in self.linhas)

    @property
    def iloc(self) -> _Posicional:
        return _Posicional(self)

    def itertuples(self, index: bool = False, name: str | None = "Linha"):
        """Percorre as linhas.

        Com ``name=None`` devolve tuplas simples (a ordem é a de ``colunas``);
        caso contrário, objetos com acesso por atributo — ``linha.nome``.
        """
        _ = index  # a Tabela não tem índice; o parâmetro existe por compatibilidade
        # Colunas de ano ("2025") não são identificadores Python válidos, então
        # nem toda tabela aceita acesso por atributo — nesse caso, tuplas.
        if name is None or not all(c.isidentifier() for c in self.colunas):
            for linha in self.linhas:
                yield tuple(linha.get(coluna) for coluna in self.colunas)
            return
        tipo = NamedTuple(name, [(c, Any) for c in self.colunas])  # type: ignore[misc]
        for linha in self.linhas:
            yield tipo(*(linha.get(coluna) for coluna in self.colunas))

    def to_dict(self, orient: str = "records") -> list[dict]:
        if orient != "records":
            raise ValueError(f"orient não suportado: {orient}")
        return [dict(linha) for linha in self.linhas]

    # -- transformação ------------------------------------------------------

    def copy(self) -> Tabela:
        return Tabela([dict(linha) for linha in self.linhas], self.colunas)

    def head(self, quantidade: int) -> Tabela:
        return Tabela(self.linhas[:quantidade], self.colunas)

    def sort_values(self, por, ascending=True) -> Tabela:
        """Ordena por uma coluna ou por várias (``ascending`` pode ser lista).

        Nulos viram string vazia para não estourar comparando com texto —
        é o que o pandas fazia com ``na_position='first'``.
        """
        colunas = [por] if isinstance(por, str) else list(por)
        crescentes = (
            [ascending] * len(colunas) if isinstance(ascending, bool) else list(ascending)
        )
        linhas = list(self.linhas)
        # Ordena da última chave para a primeira: assim a primeira predomina
        # (ordenação estável do Python).
        for coluna, crescente in reversed(list(zip(colunas, crescentes))):
            linhas.sort(key=lambda li, c=coluna: _chave_ordem(li.get(c)),
                        reverse=not crescente)
        return Tabela(linhas, self.colunas)

    def definir(self, mascara: Mascara, coluna: str, valor: Any) -> None:
        """Grava ``valor`` na ``coluna`` das linhas marcadas (era ``.loc[]``)."""
        for linha, marcada in zip(self.linhas, mascara):
            if marcada:
                linha[coluna] = valor


class _Posicional:
    """Acesso por posição: ``tabela.iloc[0]`` devolve a linha como dicionário."""

    __slots__ = ("tabela",)

    def __init__(self, tabela: Tabela) -> None:
        self.tabela = tabela

    def __getitem__(self, indice: int) -> dict:
        return self.tabela.linhas[indice]


def _chave_ordem(valor: Any) -> tuple[int, Any]:
    """Ordena números antes de texto, e nulos antes de tudo."""
    if valor is None:
        return (0, "")
    if isinstance(valor, (int, float)):
        return (1, valor)
    return (2, str(valor))


def filtrar(tabela: Tabela, termo: str, colunas: list[str]) -> Tabela:
    """Filtra por texto livre nas colunas indicadas (busca das telas)."""
    termo = termo.strip().lower()
    if not termo:
        return tabela.copy()
    presentes = [c for c in colunas if c in tabela.colunas]
    if not presentes:
        return Tabela([], tabela.colunas)
    return Tabela(
        [
            linha
            for linha in tabela.linhas
            if any(
                termo in ("" if linha.get(c) is None else str(linha.get(c))).lower()
                for c in presentes
            )
        ],
        tabela.colunas,
    )


# ---------------------------------------------------------------------------
# Leitura de planilhas (.xlsx) — era pd.read_excel
# ---------------------------------------------------------------------------


def _normalizar_celula(valor: Any) -> Any:
    """Número inteiro escrito como float vira int — como o pandas entregava.

    O calamine devolve toda célula numérica como float, então o ano do
    cabeçalho da planilha de temas chegava como ``2024.0``; o código que lê
    isso faz ``int(str(valor))``, que engasga com o ".0" e descartava a coluna
    do ano inteira.
    """
    if isinstance(valor, float) and valor.is_integer():
        return int(valor)
    return valor


class Grade:
    """Células de uma aba por posição: ``grade.iloc[linha, coluna]``.

    A planilha de temas (S-99) não tem cabeçalho fixo — o código procura a
    linha "ANO:" e lê por índice —, então o que serve aqui é uma grade crua,
    não uma tabela com colunas nomeadas.
    """

    __slots__ = ("celulas",)

    def __init__(self, celulas: list[list[Any]]) -> None:
        self.celulas = celulas

    def __len__(self) -> int:
        return len(self.celulas)

    @property
    def shape(self) -> tuple[int, int]:
        return (len(self.celulas), max((len(li) for li in self.celulas), default=0))

    @property
    def iloc(self) -> _GradePosicional:
        return _GradePosicional(self)


class _GradePosicional:
    __slots__ = ("grade",)

    def __init__(self, grade: Grade) -> None:
        self.grade = grade

    def __getitem__(self, pos: tuple[int, int]) -> Any:
        linha, coluna = pos
        if 0 <= linha < len(self.grade.celulas):
            celulas = self.grade.celulas[linha]
            if 0 <= coluna < len(celulas):
                return celulas[coluna]
        return None


def ler_planilha_crua(caminho) -> Grade:
    """Lê a primeira aba de um .xlsx como grade de células, sem cabeçalho.

    Tenta o python-calamine primeiro: o Temas.xlsx oficial (S-99) tem um
    stylesheet que o openpyxl recusa a abrir ("Max value is 14"), e era por
    isso que a versão com pandas também pedia esse engine. O openpyxl fica
    como reserva para onde o calamine não existe (Android), onde de todo modo
    a carga vem do temas_seed.json.
    """
    try:
        from python_calamine import CalamineWorkbook

        livro = CalamineWorkbook.from_path(str(caminho))
        return Grade(
            [
                [_normalizar_celula(c) for c in linha]
                for linha in livro.get_sheet_by_index(0).to_python()
            ]
        )
    except ImportError:
        pass

    from openpyxl import load_workbook

    livro_xl = load_workbook(caminho, data_only=True, read_only=True)
    try:
        aba = livro_xl[livro_xl.sheetnames[0]]
        return Grade([list(linha) for linha in aba.iter_rows(values_only=True)])
    finally:
        livro_xl.close()


def ler_aba_com_cabecalho(caminho, aba: str) -> list[dict] | None:
    """Lê uma aba pelo nome, usando a 1ª linha como cabeçalho.

    Devolve ``list[dict]`` com os valores como texto (era ``dtype=str``), ou
    ``None`` se a aba não existir — quem chama distingue "aba ausente" de
    "arquivo inválido".
    """
    from openpyxl import load_workbook

    livro = load_workbook(caminho, data_only=True, read_only=True)
    try:
        if aba not in livro.sheetnames:
            return None
        linhas = list(livro[aba].iter_rows(values_only=True))
    finally:
        livro.close()

    if not linhas:
        return []
    cabecalhos = ["" if c is None else str(c).strip() for c in linhas[0]]
    registros = []
    for linha in linhas[1:]:
        registro = {}
        for indice, cabecalho in enumerate(cabecalhos):
            if not cabecalho:
                continue
            valor = linha[indice] if indice < len(linha) else None
            registro[cabecalho] = None if valor is None else str(valor)
        registros.append(registro)
    return registros
