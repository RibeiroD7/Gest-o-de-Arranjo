"""Planilha-modelo para carga inicial de dados do aplicativo.

Gera um arquivo .xlsx com abas de Congregações, Oradores e Presidentes para a
congregação preencher, e importa a planilha preenchida para o banco. Os temas
ficam de fora: são importados diretamente dos formulários S-99/S-99a na aba
Temas.

A importação cria o que falta e completa o que já existe: para um registro
já cadastrado, só as colunas preenchidas na planilha são gravadas por cima.
Coluna em branco não apaga nada.
"""

import re
import sqlite3
from datetime import datetime

from armazenamento import EXPORTS_DIR
from database import get_connection

ABA_LEIA_ME = "Leia-me"
ABA_CONGREGACOES = "Congregações"
ABA_ORADORES = "Oradores"
ABA_PRESIDENTES = "Presidentes"

COLUNAS_CONGREGACOES = [
    "Nome*", "Responsável", "Telefone", "Endereço",
    "Dia da reunião", "Horário", "Observações",
]
COLUNAS_ORADORES = [
    "Nome*", "Telefone", "Categoria", "Congregação",
    "Temas que faz (números separados por vírgula)", "Observações",
]
COLUNAS_PRESIDENTES = ["Nome*", "Categoria"]

CATEGORIAS_VALIDAS = ("Ancião", "Servo Ministerial")

# Título provisório para temas citados na planilha antes do S-99 ser importado;
# a importação do S-99 na aba Temas substitui pelo título oficial.
TITULO_TEMA_PENDENTE = "(título pendente: importe o S-99 na aba Temas)"

INSTRUCOES = [
    "PLANILHA DE DADOS — GESTÃO DE ARRANJO",
    "",
    "Preencha as abas desta planilha e importe o arquivo em Ajustes → "
    "Importar planilha preenchida. Colunas marcadas com * são obrigatórias; "
    "as demais podem ficar em branco.",
    "",
    "A importação CRIA o que falta e COMPLETA o que já existe: para quem já "
    "está cadastrado, só as colunas preenchidas aqui são gravadas por cima. "
    "Coluna em branco não apaga nada.",
    "",
    "Aba Congregações: uma linha por congregação.",
    "   Exemplo: Jardim Primavera | João Silva | (11) 91234-5678 | "
    "Rua das Flores, 100 | Sábado | 19:30 |",
    "",
    "Aba Oradores — uma linha por orador.",
    "   • Categoria: Ancião ou Servo Ministerial (há lista suspensa na célula).",
    "   • Congregação: escreva o nome exatamente como na aba Congregações. "
    "Congregações citadas aqui e ainda não cadastradas serão criadas "
    "automaticamente.",
    "   • Temas que faz: os números dos discursos, separados por vírgula. "
    "Exemplo: 1, 5, 23, 110. Se o orador faz qualquer tema, deixe em branco "
    "e escreva \"Qualquer tema\" em Observações.",
    "",
    "Aba Presidentes: quem pode presidir a reunião de fim de semana.",
    "",
    "Os temas (títulos dos discursos) NÃO entram nesta planilha: importe os "
    "formulários oficiais S-99/S-99a em PDF diretamente na aba Temas.",
]


def gerar_planilha_modelo() -> str:
    """Cria a planilha-modelo em `exports/`; retorna o caminho absoluto."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()

    ws_leia = wb.active
    ws_leia.title = ABA_LEIA_ME
    ws_leia.column_dimensions["A"].width = 100
    for i, linha in enumerate(INSTRUCOES, start=1):
        celula = ws_leia.cell(row=i, column=1, value=linha)
        celula.alignment = Alignment(wrap_text=True, vertical="top")
        if i == 1:
            celula.font = Font(bold=True, size=13)

    fonte_cabecalho = Font(bold=True, color="FFFFFF")
    fundo_cabecalho = PatternFill("solid", fgColor="1F6E62")

    def criar_aba(titulo: str, colunas: list[str], larguras: list[int]) -> None:
        ws = wb.create_sheet(titulo)
        for idx, (nome, largura) in enumerate(zip(colunas, larguras), start=1):
            celula = ws.cell(row=1, column=idx, value=nome)
            celula.font = fonte_cabecalho
            celula.fill = fundo_cabecalho
            ws.column_dimensions[get_column_letter(idx)].width = largura
        ws.freeze_panes = "A2"
        if "Categoria" in colunas:
            col = get_column_letter(colunas.index("Categoria") + 1)
            validacao = DataValidation(
                type="list",
                formula1=f'"{",".join(CATEGORIAS_VALIDAS)}"',
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Categoria inválida",
                error="Use Ancião ou Servo Ministerial.",
            )
            ws.add_data_validation(validacao)
            validacao.add(f"{col}2:{col}300")

    criar_aba(ABA_CONGREGACOES, COLUNAS_CONGREGACOES, [30, 24, 18, 40, 16, 10, 30])
    criar_aba(ABA_ORADORES, COLUNAS_ORADORES, [30, 18, 18, 30, 44, 30])
    criar_aba(ABA_PRESIDENTES, COLUNAS_PRESIDENTES, [30, 18])

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    caminho = EXPORTS_DIR / f"planilha_dados_gestao_arranjo_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    wb.save(caminho)
    return str(caminho.resolve())


def _texto(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor != valor:  # NaN
        return ""
    return str(valor).strip()


def _normalizar_categoria(valor: str, avisos: list[str], contexto: str) -> str | None:
    texto = _texto(valor)
    if not texto:
        return None
    minusculas = texto.lower()
    if minusculas.startswith("anci"):
        return "Ancião"
    if minusculas.startswith("servo"):
        return "Servo Ministerial"
    avisos.append(f"{contexto}: categoria \"{texto}\" não reconhecida — ficou em branco.")
    return None


def _extrair_numeros_temas(valor, avisos: list[str], contexto: str) -> list[int]:
    texto = _texto(valor)
    if isinstance(valor, (int, float)) and not isinstance(valor, bool) and valor == valor:
        texto = str(int(valor))
    if not texto:
        return []
    numeros: list[int] = []
    for pedaco in re.split(r"[,;/\s]+", texto):
        if not pedaco:
            continue
        try:
            numeros.append(int(pedaco))
        except ValueError:
            avisos.append(f"{contexto}: \"{pedaco}\" não é um número de tema — ignorado.")
    return sorted(set(numeros))


def _ler_aba(caminho: str, aba: str):
    """Lê uma aba da planilha; retorna list[dict] ou None se a aba não existir."""
    from tabela import ler_aba_com_cabecalho

    return ler_aba_com_cabecalho(caminho, aba)


# Colunas da planilha e a coluna correspondente no banco. Célula vazia não
# apaga o que já está lá: quem preenche a planilha costuma preencher só o que
# falta, e uma coluna em branco significa "não mexa", não "apague".
CAMPOS_CONGREGACAO = (
    ("Responsável", "responsavel"),
    ("Telefone", "telefone"),
    ("Endereço", "endereco"),
    ("Dia da reunião", "dia_semana"),
    ("Horário", "horario"),
    ("Observações", "observacoes"),
)
CAMPOS_ORADOR = (
    ("Telefone", "telefone"),
    ("Observações", "observacoes"),
)


def _mudancas(row: dict, atual: dict, campos) -> dict:
    """O que a planilha muda neste registro: só coluna preenchida e diferente."""
    novos = {}
    for rotulo, coluna in campos:
        valor = _texto(row.get(rotulo))
        if valor and valor != _texto(atual.get(coluna)):
            novos[coluna] = valor
    return novos


def _gravar_mudancas(cursor, tabela: str, registro_id: int, mudancas: dict) -> None:
    atribuicoes = ", ".join(f"{coluna} = ?" for coluna in mudancas)
    cursor.execute(
        f"UPDATE {tabela} SET {atribuicoes} WHERE id = ?",  # noqa: S608 — colunas fixas
        (*mudancas.values(), registro_id),
    )


def importar_planilha_dados(caminho: str) -> dict:
    """Importa a planilha preenchida: cria o que falta e completa o que existe.

    Nada é apagado. Para quem já está cadastrado, só as colunas preenchidas na
    planilha são gravadas por cima — é assim que dá para exportar a lista do
    que está incompleto, preencher no Excel e trazer de volta.

    Retorna {"congregacoes": {...}, "oradores": {...}, "presidentes": {...},
    "avisos": [...]}, cada um com novas/novos, atualizadas/atualizados e
    sem_mudanca.
    """
    avisos: list[str] = []
    resumo = {
        "congregacoes": {"novas": 0, "atualizadas": 0, "sem_mudanca": 0},
        "oradores": {"novos": 0, "atualizados": 0, "sem_mudanca": 0},
        "presidentes": {"novos": 0, "atualizados": 0, "sem_mudanca": 0},
        "avisos": avisos,
    }

    try:
        df_cong = _ler_aba(caminho, ABA_CONGREGACOES)
        df_orad = _ler_aba(caminho, ABA_ORADORES)
        df_pres = _ler_aba(caminho, ABA_PRESIDENTES)
    except Exception as exc:
        raise ValueError("Não foi possível ler o arquivo como planilha .xlsx.") from exc
    if df_cong is None and df_orad is None and df_pres is None:
        raise ValueError(
            "A planilha não tem as abas esperadas (Congregações, Oradores, "
            "Presidentes). Use a planilha-modelo baixada em Ajustes."
        )

    conn = get_connection()
    try:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cadastradas = {
            dict(linha)["nome"].strip().lower(): dict(linha)
            for linha in cursor.execute("SELECT * FROM congregacoes")
        }

        def obter_congregacao(nome: str) -> int | None:
            """Retorna o id da congregação, criando-a se ainda não existir."""
            chave = nome.strip().lower()
            if not chave:
                return None
            if chave not in cadastradas:
                cursor.execute(
                    "INSERT INTO congregacoes (nome, responsavel, telefone, endereco, "
                    "dia_semana, horario, observacoes) VALUES (?, '', '', '', '', '', '')",
                    (nome.strip(),),
                )
                cadastradas[chave] = {"id": int(cursor.lastrowid), "nome": nome.strip()}
                resumo["congregacoes"]["novas"] += 1
            return cadastradas[chave]["id"]

        if df_cong is not None:
            for row in df_cong:
                nome = _texto(row.get("Nome*"))
                if not nome:
                    continue
                chave = nome.lower()
                if chave in cadastradas:
                    atual = cadastradas[chave]
                    mudancas = _mudancas(row, atual, CAMPOS_CONGREGACAO)
                    if mudancas:
                        _gravar_mudancas(cursor, "congregacoes", atual["id"], mudancas)
                        atual.update(mudancas)
                        resumo["congregacoes"]["atualizadas"] += 1
                    else:
                        resumo["congregacoes"]["sem_mudanca"] += 1
                    continue
                cursor.execute(
                    """
                    INSERT INTO congregacoes (
                        nome, responsavel, telefone, endereco, dia_semana, horario, observacoes
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        nome,
                        _texto(row.get("Responsável")),
                        _texto(row.get("Telefone")),
                        _texto(row.get("Endereço")),
                        _texto(row.get("Dia da reunião")),
                        _texto(row.get("Horário")),
                        _texto(row.get("Observações")),
                    ),
                )
                cadastradas[chave] = {
                    "id": int(cursor.lastrowid),
                    "nome": nome,
                    "responsavel": _texto(row.get("Responsável")),
                    "telefone": _texto(row.get("Telefone")),
                    "endereco": _texto(row.get("Endereço")),
                    "dia_semana": _texto(row.get("Dia da reunião")),
                    "horario": _texto(row.get("Horário")),
                    "observacoes": _texto(row.get("Observações")),
                }
                resumo["congregacoes"]["novas"] += 1

        if df_orad is not None:
            existentes = {
                (dict(linha)["nome"].strip().lower(), dict(linha)["congregacao_id"]): dict(linha)
                for linha in cursor.execute(
                    "SELECT id, nome, telefone, categoria, congregacao_id, observacoes FROM oradores"
                )
            }
            temas_existentes = {int(nr) for (nr,) in cursor.execute("SELECT nr FROM temas")}
            temas_pendentes_criados = False

            def garantir_tema(nr: int) -> None:
                """Cria um tema provisório para o vínculo orador↔tema não se perder."""
                nonlocal temas_pendentes_criados
                if nr in temas_existentes:
                    return
                cursor.execute(
                    "INSERT INTO temas (nr, titulo) VALUES (?, ?)",
                    (nr, TITULO_TEMA_PENDENTE),
                )
                temas_existentes.add(nr)
                temas_pendentes_criados = True

            def vincular_temas(orador_id: int, row: dict, contexto: str) -> bool:
                """Liga o orador aos temas citados. Só acrescenta; nunca desliga."""
                mudou = False
                for tema_nr in _extrair_numeros_temas(
                    row.get("Temas que faz (números separados por vírgula)"), avisos, contexto
                ):
                    garantir_tema(tema_nr)
                    cursor.execute(
                        "INSERT OR IGNORE INTO orador_temas (orador_id, tema_nr) VALUES (?, ?)",
                        (orador_id, tema_nr),
                    )
                    mudou = mudou or cursor.rowcount > 0
                return mudou

            for i, row in enumerate(df_orad):
                nome = _texto(row.get("Nome*"))
                if not nome:
                    continue
                contexto = f"Oradores, linha {i + 2}"
                congregacao_id = obter_congregacao(_texto(row.get("Congregação")))
                categoria = _normalizar_categoria(row.get("Categoria"), avisos, contexto)
                atual = existentes.get((nome.lower(), congregacao_id))
                if atual:
                    mudancas = _mudancas(row, atual, CAMPOS_ORADOR)
                    if categoria and categoria != _texto(atual.get("categoria")):
                        mudancas["categoria"] = categoria
                    if mudancas:
                        _gravar_mudancas(cursor, "oradores", atual["id"], mudancas)
                        atual.update(mudancas)
                    ligou_tema = vincular_temas(atual["id"], row, contexto)
                    if mudancas or ligou_tema:
                        resumo["oradores"]["atualizados"] += 1
                    else:
                        resumo["oradores"]["sem_mudanca"] += 1
                    continue
                cursor.execute(
                    """
                    INSERT INTO oradores (nome, telefone, categoria, congregacao_id, observacoes)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (
                        nome,
                        _texto(row.get("Telefone")),
                        categoria,
                        congregacao_id,
                        _texto(row.get("Observações")),
                    ),
                )
                orador_id = int(cursor.lastrowid)
                existentes[(nome.lower(), congregacao_id)] = {
                    "id": orador_id,
                    "nome": nome,
                    "telefone": _texto(row.get("Telefone")),
                    "categoria": categoria,
                    "congregacao_id": congregacao_id,
                    "observacoes": _texto(row.get("Observações")),
                }
                resumo["oradores"]["novos"] += 1
                vincular_temas(orador_id, row, contexto)

            if temas_pendentes_criados:
                avisos.append(
                    "Alguns temas citados na planilha ainda não estão cadastrados; "
                    "foram criados com título pendente. Importe o formulário S-99 "
                    "na aba Temas para preencher os títulos oficiais."
                )

        if df_pres is not None:
            presidentes = {
                dict(linha)["nome"].strip().lower(): dict(linha)
                for linha in cursor.execute(
                    "SELECT id, nome, categoria FROM presidentes_cadastro"
                )
            }
            for i, row in enumerate(df_pres):
                nome = _texto(row.get("Nome*"))
                if not nome:
                    continue
                categoria = _normalizar_categoria(
                    row.get("Categoria"), avisos, f"Presidentes, linha {i + 2}"
                )
                atual = presidentes.get(nome.lower())
                if atual:
                    if categoria and categoria != _texto(atual.get("categoria")):
                        _gravar_mudancas(
                            cursor, "presidentes_cadastro", atual["id"], {"categoria": categoria}
                        )
                        atual["categoria"] = categoria
                        resumo["presidentes"]["atualizados"] += 1
                    else:
                        resumo["presidentes"]["sem_mudanca"] += 1
                    continue
                cursor.execute(
                    """
                    INSERT INTO presidentes_cadastro (nome, categoria, ordem)
                    VALUES (?, ?, (SELECT COALESCE(MAX(ordem), 0) + 1 FROM presidentes_cadastro))
                    """,
                    (nome, categoria or "Ancião"),
                )
                presidentes[nome.lower()] = {
                    "id": int(cursor.lastrowid),
                    "nome": nome,
                    "categoria": categoria or "Ancião",
                }
                resumo["presidentes"]["novos"] += 1

        conn.commit()
    finally:
        conn.close()

    return resumo
