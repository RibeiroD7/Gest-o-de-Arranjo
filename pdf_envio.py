"""
Geração do PDF de envio preenchendo o template Excel models/MODELO ENVIO.xlsx.

No Windows usa automação real do Excel (win32com) para máxima fidelidade visual.
Em outros sistemas (Linux/macOS), preenche o template com openpyxl e converte
para PDF via LibreOffice (requer o comando `soffice` ou `libreoffice` instalado).
"""

from __future__ import annotations

import platform
import shutil
import subprocess
import uuid
from datetime import datetime
from pathlib import Path

from database import get_connection

EXPORTS_DIR = Path("exports")
TEMPLATE_PATH = Path("models") / "MODELO ENVIO.xlsx"
SHEET_NAME = "Oradores - Envio"
LINHA_INICIO_DADOS = 15
LINHA_FINAL_LIMPEZA = 200

XL_TYPE_PDF = 0


def _formatar_esbocos(temas: str, observacoes: str) -> str:
    if observacoes and "qualquer tema" in observacoes.lower():
        return "Qualquer tema"
    if temas and str(temas).strip():
        return str(temas).strip()
    return ""


def _formatar_nome_orador(nome: str, categoria: str) -> str:
    nome = (nome or "").strip()
    categoria = (categoria or "").strip()
    if nome and categoria:
        return f"{nome} ({categoria})"
    return nome


def _linhas_endereco(endereco: str) -> tuple[str, str]:
    if not (endereco or "").strip():
        return "", ""
    partes = [p.strip() for p in endereco.replace("\n", " - ").split(" - ") if p.strip()]
    if not partes:
        return "", ""
    if len(partes) == 1:
        return partes[0], ""
    return partes[0], partes[1]


def _carregar_configuracao() -> dict:
    conn = get_connection()
    try:
        cursor = conn.execute(
            """
            SELECT nome_congregacao, endereco, cidade, cep, coordenador_discursos,
                   telefone_coordenador, dia_reuniao, horario_reuniao, circuito
            FROM configuracoes WHERE id = 1
            """
        )
        row = cursor.fetchone()
    finally:
        conn.close()

    if not row:
        return {
            "nome_congregacao": "",
            "endereco": "",
            "cidade": "",
            "cep": "",
            "coordenador_discursos": "",
            "telefone_coordenador": "",
            "dia_reuniao": "",
            "horario_reuniao": "",
            "circuito": "",
        }

    return {
        "nome_congregacao": row[0] or "",
        "endereco": row[1] or "",
        "cidade": row[2] or "",
        "cep": row[3] or "",
        "coordenador_discursos": row[4] or "",
        "telefone_coordenador": row[5] or "",
        "dia_reuniao": row[6] or "",
        "horario_reuniao": row[7] or "",
        "circuito": row[8] or "",
    }


def carregar_oradores_por_ids(orador_ids: list[int]) -> list[dict]:
    if not orador_ids:
        return []

    placeholders = ",".join("?" * len(orador_ids))
    conn = get_connection()
    try:
        cursor = conn.execute(
            f"""
            SELECT o.id,
                   o.nome,
                   o.telefone,
                   o.categoria,
                   o.observacoes,
                   COALESCE((
                       SELECT GROUP_CONCAT(ot.tema_nr, ', ')
                       FROM orador_temas ot
                       WHERE ot.orador_id = o.id
                   ), '') AS temas
            FROM oradores o
            WHERE o.id IN ({placeholders})
            ORDER BY o.categoria, o.nome
            """,
            orador_ids,
        )
        colunas = [desc[0] for desc in cursor.description]
        return [dict(zip(colunas, linha)) for linha in cursor.fetchall()]
    finally:
        conn.close()


def _nome_arquivo_pdf() -> str:
    return f"Lista_Oradores_Envio_{datetime.now().strftime('%Y-%m')}.pdf"


def _linha_reuniao(dia: str, horario: str) -> str:
    if dia and horario:
        return f"{dia}, {horario}"
    return dia or horario or ""


def _preencher_cabecalho_win32(ws, config: dict) -> None:
    nome = config.get("nome_congregacao") or "Minha congregação"
    cidade = (config.get("cidade") or "").strip()
    cep = (config.get("cep") or "").strip()
    circuito = (config.get("circuito") or "").strip()
    coordenador = (config.get("coordenador_discursos") or "").strip()
    telefone = (config.get("telefone_coordenador") or "").strip()
    reuniao = _linha_reuniao(
        (config.get("dia_reuniao") or "").strip(),
        (config.get("horario_reuniao") or "").strip(),
    )
    endereco_linha1, endereco_linha2 = _linhas_endereco(config.get("endereco", ""))

    ws.Range("C1").Value = f"{nome} - {cidade}"
    ws.Range("A3").Value = f"{nome} ({circuito})"
    ws.Range("C3").Value = "Coordenador de discursos públicos"
    ws.Range("A4").Value = endereco_linha1
    ws.Range("C4").Value = coordenador
    ws.Range("A5").Value = endereco_linha2
    ws.Range("C5").Value = f"Tel {telefone}" if telefone else ""
    ws.Range("A6").Value = cidade
    ws.Range("A7").Value = cep
    ws.Range("A10").Value = "Reunião de fim de semana "
    ws.Range("A11").Value = reuniao


def _preencher_tabela_win32(ws, oradores: list[dict]) -> int:
    ws.Range(f"A{LINHA_INICIO_DADOS}:D{LINHA_FINAL_LIMPEZA}").ClearContents()

    linha = LINHA_INICIO_DADOS
    for orador in oradores:
        ws.Cells(linha, 1).Value = _formatar_nome_orador(
            orador["nome"],
            orador.get("categoria", ""),
        )
        ws.Cells(linha, 2).Value = orador.get("telefone") or ""
        ws.Cells(linha, 3).Value = _formatar_esbocos(
            orador.get("temas", ""),
            orador.get("observacoes", ""),
        )
        ws.Cells(linha, 4).Value = orador.get("observacoes") or ""
        linha += 1

    return linha - 1


def _ajustar_area_impressao_win32(ws, ultima_linha: int) -> None:
    linha_final = max(ultima_linha + 2, 20)
    ws.PageSetup.PrintArea = f"$A$1:$D${linha_final}"


def _gerar_via_windows(
    oradores: list[dict],
    config: dict,
    pdf_path: Path,
    temp_xlsx: Path,
) -> tuple[str | None, str | None]:
    import win32com.client as win32

    excel = None
    workbook = None
    try:
        shutil.copy2(TEMPLATE_PATH, temp_xlsx)

        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False

        workbook = excel.Workbooks.Open(str(temp_xlsx))
        ws = workbook.Worksheets(SHEET_NAME)

        _preencher_cabecalho_win32(ws, config)
        ultima_linha = _preencher_tabela_win32(ws, oradores)
        _ajustar_area_impressao_win32(ws, ultima_linha)

        workbook.ExportAsFixedFormat(
            Type=XL_TYPE_PDF,
            Filename=str(pdf_path),
            Quality=0,
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
            OpenAfterPublish=False,
        )

        return str(pdf_path), None

    except Exception as exc:
        return None, f"Erro ao gerar PDF via Excel: {exc}"

    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()


def _preencher_cabecalho_openpyxl(ws, config: dict) -> None:
    nome = config.get("nome_congregacao") or "Minha congregação"
    cidade = (config.get("cidade") or "").strip()
    cep = (config.get("cep") or "").strip()
    circuito = (config.get("circuito") or "").strip()
    coordenador = (config.get("coordenador_discursos") or "").strip()
    telefone = (config.get("telefone_coordenador") or "").strip()
    reuniao = _linha_reuniao(
        (config.get("dia_reuniao") or "").strip(),
        (config.get("horario_reuniao") or "").strip(),
    )
    endereco_linha1, endereco_linha2 = _linhas_endereco(config.get("endereco", ""))

    ws["C1"] = f"{nome} - {cidade}"
    ws["A3"] = f"{nome} ({circuito})"
    ws["C3"] = "Coordenador de discursos públicos"
    ws["A4"] = endereco_linha1
    ws["C4"] = coordenador
    ws["A5"] = endereco_linha2
    ws["C5"] = f"Tel {telefone}" if telefone else ""
    ws["A6"] = cidade
    ws["A7"] = cep
    ws["A10"] = "Reunião de fim de semana "
    ws["A11"] = reuniao


def _preencher_tabela_openpyxl(ws, oradores: list[dict]) -> int:
    for linha in range(LINHA_INICIO_DADOS, LINHA_FINAL_LIMPEZA + 1):
        for coluna in range(1, 5):
            ws.cell(row=linha, column=coluna, value=None)

    linha = LINHA_INICIO_DADOS
    for orador in oradores:
        ws.cell(row=linha, column=1, value=_formatar_nome_orador(
            orador["nome"],
            orador.get("categoria", ""),
        ))
        ws.cell(row=linha, column=2, value=orador.get("telefone") or "")
        ws.cell(row=linha, column=3, value=_formatar_esbocos(
            orador.get("temas", ""),
            orador.get("observacoes", ""),
        ))
        ws.cell(row=linha, column=4, value=orador.get("observacoes") or "")
        linha += 1

    return linha - 1


def _ajustar_area_impressao_openpyxl(ws, ultima_linha: int) -> None:
    linha_final = max(ultima_linha + 2, 20)
    ws.print_area = f"A1:D{linha_final}"


def _comando_libreoffice() -> list[str] | None:
    """Retorna o prefixo de comando para invocar o LibreOffice headless.

    Verifica o binário `soffice`/`libreoffice` no PATH e, na ausência dele,
    uma instalação via Flatpak (comum em distros como Fedora Silverblue).
    """
    binario = shutil.which("soffice") or shutil.which("libreoffice")
    if binario:
        return [binario]

    if shutil.which("flatpak"):
        verificacao = subprocess.run(
            ["flatpak", "info", "org.libreoffice.LibreOffice"],
            capture_output=True,
            check=False,
        )
        if verificacao.returncode == 0:
            return ["flatpak", "run", "org.libreoffice.LibreOffice"]

    return None


def _gerar_via_libreoffice(
    oradores: list[dict],
    config: dict,
    pdf_path: Path,
    temp_xlsx: Path,
) -> tuple[str | None, str | None]:
    comando = _comando_libreoffice()
    if not comando:
        return None, (
            "Geração de PDF requer o LibreOffice instalado (binário 'soffice'/"
            "'libreoffice' ou o Flatpak 'org.libreoffice.LibreOffice') neste sistema."
        )

    import openpyxl

    shutil.copy2(TEMPLATE_PATH, temp_xlsx)
    workbook = openpyxl.load_workbook(temp_xlsx)
    ws = workbook[SHEET_NAME]

    _preencher_cabecalho_openpyxl(ws, config)
    ultima_linha = _preencher_tabela_openpyxl(ws, oradores)
    _ajustar_area_impressao_openpyxl(ws, ultima_linha)
    workbook.save(temp_xlsx)

    try:
        resultado = subprocess.run(
            [
                *comando,
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(EXPORTS_DIR.resolve()),
                str(temp_xlsx),
            ],
            capture_output=True,
            text=True,
            timeout=60,
        )
    except subprocess.TimeoutExpired:
        return None, "Tempo esgotado ao converter o arquivo para PDF via LibreOffice."

    if resultado.returncode != 0:
        return None, f"Erro ao converter para PDF via LibreOffice: {resultado.stderr.strip()}"

    gerado = temp_xlsx.with_suffix(".pdf")
    if not gerado.exists():
        return None, "LibreOffice não gerou o arquivo PDF esperado."

    gerado.replace(pdf_path)
    return str(pdf_path), None


def gerar_pdf_envio(oradores_selecionados: list[int]) -> tuple[str | None, str | None]:
    """
    Gera o PDF de envio preenchendo o template Excel.

    No Windows usa automação do Excel (win32com); nos demais sistemas usa
    openpyxl + LibreOffice headless. Retorna (caminho_arquivo, mensagem_erro).
    """
    if not oradores_selecionados:
        return None, "Nenhum orador selecionado."

    if not TEMPLATE_PATH.exists():
        return None, f"Template não encontrado: {TEMPLATE_PATH}"

    oradores = carregar_oradores_por_ids(oradores_selecionados)
    if not oradores:
        return None, "Nenhum orador válido encontrado para gerar o PDF."

    config = _carregar_configuracao()
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    pdf_path = (EXPORTS_DIR / _nome_arquivo_pdf()).resolve()
    temp_xlsx = (EXPORTS_DIR / f"_temp_envio_{uuid.uuid4().hex}.xlsx").resolve()

    try:
        if platform.system() == "Windows":
            return _gerar_via_windows(oradores, config, pdf_path, temp_xlsx)
        return _gerar_via_libreoffice(oradores, config, pdf_path, temp_xlsx)
    finally:
        if temp_xlsx.exists():
            try:
                temp_xlsx.unlink()
            except OSError:
                pass
